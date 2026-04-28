import openpyxl
from collections import defaultdict
import json

EXCEL = 'FINAL Amazon sales Backup.xlsx'
OUTPUT = '../Reportes/dashboard_mensual.html'

wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)

# ── Product family mapping ──────────────────────────────────────────
def product_family(name):
    if not name or name == 'Unknown':
        return 'Otros'
    n = name.lower()
    if 'audio calm night' in n: return 'Audio Calm Night'
    if 'audio calm' in n: return 'Audio Calm'
    if 'bacticure' in n: return 'Bacticure'
    if 'bacterium' in n: return 'Bacterium'
    if 'bee flex' in n: return 'Bee Flex'
    if 'calvistop' in n: return 'CalviStop'
    if 'collagen' in n and 'inlife' not in n: return 'Collagen'
    if 'fertili' in n: return 'Fertili-T'
    if 'fungitrim' in n: return 'FungiTrim'
    if 'hemokin' in n: return 'Hemokin'
    if 'hepatol' in n: return 'Hepatol'
    if 'inlife' in n: return 'Inlife'
    if 'lovv' in n: return 'Lovv'
    if 'probiotik' in n: return 'Probiotik'
    if 'regemax' in n: return 'Regemax'
    if 'ringoff' in n or 'ring off' in n: return 'RingOff'
    if 'roncout' in n: return 'RoncOut'
    if 'sauna' in n: return 'Sauna Twin Gel'
    if 'shilajit' in n: return 'Shilajit'
    if 'stem cell' in n: return 'Stem Cell'
    if 'vitatonic' in n: return 'Vitatonic'
    if 'vitiligo orange cream + caps' in n: return 'Vitiligo Bundle'
    if 'vitiligo orange' in n: return 'Vitiligo Orange'
    if 'vitiligo purple' in n: return 'Vitiligo Purple'
    if 'vitiligo' in n: return 'Vitiligo'
    return 'Otros'


# ── Extract ventas_amz ──────────────────────────────────────────────
ws = wb['ventas_amz']
monthly = defaultdict(lambda: {
    'qty': 0, 'sales': 0, 'profit': 0, 'cogs': 0,
    'amz_fee': 0, 'weeks': set()
})
by_product = defaultdict(lambda: defaultdict(lambda: {
    'qty': 0, 'sales': 0, 'profit': 0
}))

for i, r in enumerate(ws.iter_rows(values_only=True)):
    if i < 2:
        continue
    row = list(r)
    status = row[7]
    retailer = row[9]
    if status in ('Cancelled', 'Pending'):
        continue
    if retailer != 'Amazon.com':
        continue

    bcm = row[55]
    bcw = row[54]
    if not bcm:
        continue

    qty = row[17] or 0
    sales = row[19] or 0
    profit = row[32] or 0
    cogs = row[30] or 0
    amz_fee = row[29] or 0
    prod = product_family(row[33])

    monthly[bcm]['qty'] += qty
    monthly[bcm]['sales'] += sales
    monthly[bcm]['profit'] += profit
    monthly[bcm]['cogs'] += cogs
    monthly[bcm]['amz_fee'] += amz_fee
    if bcw:
        monthly[bcm]['weeks'].add(bcw)

    by_product[prod][bcm]['qty'] += qty
    by_product[prod][bcm]['sales'] += sales
    by_product[prod][bcm]['profit'] += profit


# ── Extract PPC_Amz ─────────────────────────────────────────────────
ws = wb['PPC_Amz']
ppc_monthly = defaultdict(lambda: {
    'spend': 0, 'sales': 0, 'impressions': 0, 'clicks': 0
})
ppc_by_product = defaultdict(lambda: defaultdict(lambda: {
    'spend': 0, 'sales': 0
}))

PPC_PRODUCT_MAP = {
    'Audio Calm': 'Audio Calm',
    'Bacterium': 'Bacterium',
    'Bacticure': 'Bacticure',
    'Bee Flex': 'Bee Flex',
    'CalviStop': 'CalviStop',
    'Celulas Madres': 'Stem Cell',
    'Colageno': 'Collagen',
    'FertiliT': 'Fertili-T',
    'FungiTrim': 'FungiTrim',
    'Hepatol': 'Hepatol',
    'Inlife': 'Inlife',
    'Lovv': 'Lovv',
    'Probiotik': 'Probiotik',
    'Regemax': 'Regemax',
    'Ring Off': 'RingOff',
    'RoncOut': 'RoncOut',
    'SPF': 'Otros',
    'Shilajit': 'Shilajit',
    'Vitatonic': 'Vitatonic',
    'Vitiligo': 'Vitiligo',
}

for i, r in enumerate(ws.iter_rows(values_only=True)):
    if i < 2:
        continue
    row = list(r)
    bcm = row[32]
    if not bcm:
        continue

    spend = row[16] or 0
    ppc_sales = row[17] or 0
    impressions = row[12] or 0
    clicks = row[13] or 0
    ppc_prod = PPC_PRODUCT_MAP.get(row[29], 'Otros')

    ppc_monthly[bcm]['spend'] += spend
    ppc_monthly[bcm]['sales'] += ppc_sales
    ppc_monthly[bcm]['impressions'] += impressions
    ppc_monthly[bcm]['clicks'] += clicks

    ppc_by_product[ppc_prod][bcm]['spend'] += spend
    ppc_by_product[ppc_prod][bcm]['sales'] += ppc_sales

wb.close()


# ── Build consolidated data ─────────────────────────────────────────
all_months = sorted(set(list(monthly.keys()) + list(ppc_monthly.keys())))
# Exclude months with very little data (< 2 weeks)
all_months = [m for m in all_months if len(monthly[m]['weeks']) >= 2]

months_labels = []
data = {
    'qweeks': [], 'qty': [], 'gsales': [], 'profit': [], 'cogs': [],
    'amz_fee': [], 'ppc_sales': [], 'ppc_spend': [], 'impressions': [],
    'clicks': [], 'roas': [], 'acos': [], 'tacos': [],
    'cpc_vs_total': [], 'real_net_profit': [],
    'profit_per_week': [], 'sales_per_week': [],
    'dollar_weekly_pct': [], 'qty_weekly_pct': [],
    'profit_margin': [],
}

prev_profit_pw = None
prev_qty_pw = None

for m in all_months:
    v = monthly[m]
    p = ppc_monthly[m]
    qw = len(v['weeks']) or 1

    gsales = v['sales']
    profit = v['profit']
    ppc_spend = p['spend']
    ppc_sales = p['sales']
    roas = ppc_sales / ppc_spend if ppc_spend > 0 else 0
    acos = (ppc_spend / ppc_sales * 100) if ppc_sales > 0 else 0
    tacos = (ppc_spend / gsales * 100) if gsales > 0 else 0
    cpc_vs = (ppc_sales / gsales * 100) if gsales > 0 else 0
    real_net = profit - ppc_spend
    ppw = real_net / qw
    spw = v['qty'] / qw
    margin = (profit / gsales * 100) if gsales > 0 else 0

    d_pct = ((ppw - prev_profit_pw) / prev_profit_pw * 100) if prev_profit_pw and prev_profit_pw != 0 else None
    q_pct = ((spw - prev_qty_pw) / prev_qty_pw * 100) if prev_qty_pw and prev_qty_pw != 0 else None

    months_labels.append(m)
    data['qweeks'].append(qw)
    data['qty'].append(round(v['qty']))
    data['gsales'].append(round(gsales, 2))
    data['profit'].append(round(profit, 2))
    data['cogs'].append(round(v['cogs'], 2))
    data['amz_fee'].append(round(v['amz_fee'], 2))
    data['ppc_sales'].append(round(ppc_sales, 2))
    data['ppc_spend'].append(round(ppc_spend, 2))
    data['impressions'].append(round(p['impressions']))
    data['clicks'].append(round(p['clicks']))
    data['roas'].append(round(roas, 2))
    data['acos'].append(round(acos, 2))
    data['tacos'].append(round(tacos, 2))
    data['cpc_vs_total'].append(round(cpc_vs, 2))
    data['real_net_profit'].append(round(real_net, 2))
    data['profit_per_week'].append(round(ppw, 2))
    data['sales_per_week'].append(round(spw, 1))
    data['dollar_weekly_pct'].append(round(d_pct, 2) if d_pct is not None else None)
    data['qty_weekly_pct'].append(round(q_pct, 2) if q_pct is not None else None)
    data['profit_margin'].append(round(margin, 2))

    prev_profit_pw = ppw
    prev_qty_pw = spw


# ── Product data ────────────────────────────────────────────────────
product_families = sorted(by_product.keys())
prod_data = {}
for fam in product_families:
    pd = {'qty': [], 'sales': [], 'profit': []}
    for m in months_labels:
        d = by_product[fam].get(m, {'qty': 0, 'sales': 0, 'profit': 0})
        pd['qty'].append(round(d['qty']))
        pd['sales'].append(round(d['sales'], 2))
        pd['profit'].append(round(d['profit'], 2))
    prod_data[fam] = pd

# Top products by total sales
prod_totals = []
for fam in product_families:
    total_sales = sum(prod_data[fam]['sales'])
    total_profit = sum(prod_data[fam]['profit'])
    total_qty = sum(prod_data[fam]['qty'])
    prod_totals.append({
        'name': fam, 'sales': round(total_sales, 2),
        'profit': round(total_profit, 2), 'qty': total_qty,
        'margin': round(total_profit / total_sales * 100, 2) if total_sales > 0 else 0
    })
prod_totals.sort(key=lambda x: -x['sales'])

# PPC by product
ppc_prod_data = {}
for fam in sorted(ppc_by_product.keys()):
    pd = {'spend': [], 'sales': []}
    for m in months_labels:
        d = ppc_by_product[fam].get(m, {'spend': 0, 'sales': 0})
        pd['spend'].append(round(d['spend'], 2))
        pd['sales'].append(round(d['sales'], 2))
    ppc_prod_data[fam] = pd


# ── Summary cards data ──────────────────────────────────────────────
latest = -1
prev = -2
summary = {
    'current_month': months_labels[latest],
    'prev_month': months_labels[prev],
    'gsales_current': data['gsales'][latest],
    'gsales_prev': data['gsales'][prev],
    'profit_current': data['real_net_profit'][latest],
    'profit_prev': data['real_net_profit'][prev],
    'roas_current': data['roas'][latest],
    'roas_prev': data['roas'][prev],
    'qty_current': data['qty'][latest],
    'qty_prev': data['qty'][prev],
    'ppc_spend_current': data['ppc_spend'][latest],
    'ppc_spend_prev': data['ppc_spend'][prev],
    'tacos_current': data['tacos'][latest],
    'tacos_prev': data['tacos'][prev],
    'qweeks_current': data['qweeks'][latest],
}


# ── Generate HTML ───────────────────────────────────────────────────
html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Amazon Monthly Dashboard</title>
<script src="https://cdn.plot.ly/plotly-2.35.0.min.js"></script>
<style>
  :root {{
    --bg: #0f1117;
    --card: #1a1d29;
    --border: #2a2d3a;
    --text: #e4e4e7;
    --muted: #9ca3af;
    --green: #22c55e;
    --red: #ef4444;
    --blue: #3b82f6;
    --purple: #a855f7;
    --orange: #f97316;
    --yellow: #eab308;
    --cyan: #06b6d4;
  }}
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{
    font-family: 'Inter', -apple-system, sans-serif;
    background: var(--bg);
    color: var(--text);
    padding: 24px;
    line-height: 1.5;
  }}
  h1 {{
    font-size: 28px;
    font-weight: 700;
    margin-bottom: 4px;
  }}
  .subtitle {{
    color: var(--muted);
    font-size: 14px;
    margin-bottom: 24px;
  }}
  .cards {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 16px;
    margin-bottom: 32px;
  }}
  .card {{
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 20px;
  }}
  .card-label {{
    font-size: 12px;
    color: var(--muted);
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 8px;
  }}
  .card-value {{
    font-size: 28px;
    font-weight: 700;
  }}
  .card-change {{
    font-size: 13px;
    margin-top: 6px;
  }}
  .card-change.up {{ color: var(--green); }}
  .card-change.down {{ color: var(--red); }}
  .card-note {{
    font-size: 11px;
    color: var(--muted);
    margin-top: 4px;
  }}
  .section {{
    margin-bottom: 40px;
  }}
  .section-title {{
    font-size: 20px;
    font-weight: 600;
    margin-bottom: 16px;
    padding-bottom: 8px;
    border-bottom: 1px solid var(--border);
  }}
  .chart-row {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 20px;
    margin-bottom: 20px;
  }}
  .chart-box {{
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 16px;
  }}
  .chart-full {{
    grid-column: 1 / -1;
  }}
  .table-container {{
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 16px;
    overflow-x: auto;
  }}
  table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 13px;
  }}
  th {{
    text-align: right;
    padding: 10px 12px;
    color: var(--muted);
    font-weight: 500;
    font-size: 11px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    border-bottom: 1px solid var(--border);
  }}
  th:first-child {{ text-align: left; }}
  td {{
    text-align: right;
    padding: 10px 12px;
    border-bottom: 1px solid var(--border);
  }}
  td:first-child {{
    text-align: left;
    font-weight: 500;
  }}
  tr:last-child td {{ border-bottom: none; }}
  tr:hover td {{ background: rgba(59, 130, 246, 0.05); }}
  .pill {{
    display: inline-block;
    padding: 2px 8px;
    border-radius: 9999px;
    font-size: 11px;
    font-weight: 600;
  }}
  .pill-green {{ background: rgba(34,197,94,0.15); color: var(--green); }}
  .pill-red {{ background: rgba(239,68,68,0.15); color: var(--red); }}
  .pill-yellow {{ background: rgba(234,179,8,0.15); color: var(--yellow); }}
  .tabs {{
    display: flex;
    gap: 8px;
    margin-bottom: 16px;
    flex-wrap: wrap;
  }}
  .tab {{
    padding: 8px 16px;
    border-radius: 8px;
    border: 1px solid var(--border);
    background: transparent;
    color: var(--muted);
    cursor: pointer;
    font-size: 13px;
    transition: all 0.2s;
  }}
  .tab:hover {{ border-color: var(--blue); color: var(--text); }}
  .tab.active {{
    background: var(--blue);
    color: white;
    border-color: var(--blue);
  }}
  .plotly-chart {{ width: 100%; }}
  @media (max-width: 900px) {{
    .chart-row {{ grid-template-columns: 1fr; }}
    .cards {{ grid-template-columns: repeat(2, 1fr); }}
  }}
  .alert {{
    padding: 12px 16px;
    border-radius: 8px;
    margin-bottom: 12px;
    font-size: 13px;
  }}
  .alert-warn {{
    background: rgba(234,179,8,0.1);
    border: 1px solid rgba(234,179,8,0.3);
    color: var(--yellow);
  }}
  .alert-danger {{
    background: rgba(239,68,68,0.1);
    border: 1px solid rgba(239,68,68,0.3);
    color: var(--red);
  }}
  .alert-ok {{
    background: rgba(34,197,94,0.1);
    border: 1px solid rgba(34,197,94,0.3);
    color: var(--green);
  }}
</style>
</head>
<body>

<h1>Amazon Dashboard Mensual</h1>
<p class="subtitle">Datos hasta {months_labels[-1]} &middot; {data['qweeks'][-1]} semanas en mes actual &middot; Generado con datos del Excel</p>

<!-- ── ALERTS ────────────────────────────────────────────── -->
<div id="alerts" class="section"></div>

<!-- ── SUMMARY CARDS ─────────────────────────────────────── -->
<div class="cards" id="summary-cards"></div>

<!-- ── SALES & PROFIT ────────────────────────────────────── -->
<div class="section">
  <h2 class="section-title">Ventas & Profit</h2>
  <div class="chart-row">
    <div class="chart-box"><div id="chart-gsales"></div></div>
    <div class="chart-box"><div id="chart-profit-breakdown"></div></div>
  </div>
  <div class="chart-row">
    <div class="chart-box"><div id="chart-qty"></div></div>
    <div class="chart-box"><div id="chart-margin"></div></div>
  </div>
</div>

<!-- ── PPC / ADVERTISING ─────────────────────────────────── -->
<div class="section">
  <h2 class="section-title">PPC / Publicidad</h2>
  <div class="chart-row">
    <div class="chart-box"><div id="chart-ppc"></div></div>
    <div class="chart-box"><div id="chart-roas-acos"></div></div>
  </div>
  <div class="chart-row">
    <div class="chart-box"><div id="chart-tacos"></div></div>
    <div class="chart-box"><div id="chart-cpc-vs-total"></div></div>
  </div>
</div>

<!-- ── WEEKLY NORMALIZED ─────────────────────────────────── -->
<div class="section">
  <h2 class="section-title">Rendimiento Semanal Normalizado</h2>
  <div class="chart-row">
    <div class="chart-box"><div id="chart-weekly-profit"></div></div>
    <div class="chart-box"><div id="chart-weekly-qty"></div></div>
  </div>
  <div class="chart-row">
    <div class="chart-box"><div id="chart-weekly-pct"></div></div>
    <div class="chart-box"><div id="chart-net-profit"></div></div>
  </div>
</div>

<!-- ── PRODUCTS ───────────────────────────────────────────── -->
<div class="section">
  <h2 class="section-title">Por Producto</h2>
  <div class="tabs" id="product-tabs"></div>
  <div class="chart-row">
    <div class="chart-box chart-full"><div id="chart-product-trend"></div></div>
  </div>
  <div class="table-container" id="product-table-container"></div>
</div>

<!-- ── DATA TABLE ────────────────────────────────────────── -->
<div class="section">
  <h2 class="section-title">Tabla Resumen Mensual</h2>
  <div class="table-container" id="summary-table"></div>
</div>

<script>
const MONTHS = {json.dumps(months_labels)};
const D = {json.dumps(data)};
const SUMMARY = {json.dumps(summary)};
const PROD_DATA = {json.dumps(prod_data)};
const PROD_TOTALS = {json.dumps(prod_totals)};
const PPC_PROD = {json.dumps(ppc_prod_data)};

const plotBg = '#1a1d29';
const gridColor = '#2a2d3a';
const textColor = '#9ca3af';
const layout_base = {{
  paper_bgcolor: plotBg,
  plot_bgcolor: plotBg,
  font: {{ color: textColor, family: 'Inter, sans-serif', size: 12 }},
  margin: {{ t: 40, r: 20, b: 40, l: 60 }},
  xaxis: {{ gridcolor: gridColor, tickangle: -45 }},
  yaxis: {{ gridcolor: gridColor }},
  legend: {{ orientation: 'h', y: -0.2 }},
  hovermode: 'x unified',
}};
const config = {{ responsive: true, displayModeBar: false }};

function fmt(n) {{ return n >= 1000 ? '$' + (n/1000).toFixed(1) + 'K' : '$' + n.toFixed(0); }}
function fmtFull(n) {{ return '$' + n.toLocaleString('en-US', {{minimumFractionDigits:2, maximumFractionDigits:2}}); }}
function pct(curr, prev) {{
  if (!prev) return {{ val: 0, cls: 'up', txt: '—' }};
  let p = ((curr - prev) / Math.abs(prev)) * 100;
  return {{ val: p, cls: p >= 0 ? 'up' : 'down', txt: (p >= 0 ? '+' : '') + p.toFixed(1) + '%' }};
}}

// ── ALERTS ──────────────────────────────────────────────────
function renderAlerts() {{
  let html = '';
  let latest = D.roas[D.roas.length - 1];
  let latestAcos = D.acos[D.acos.length - 1];
  let latestTacos = D.tacos[D.tacos.length - 1];

  if (latest < 2.5) {{
    html += `<div class="alert alert-danger">ROAS actual: ${{latest.toFixed(2)}}x — por debajo del umbral 2.5x</div>`;
  }} else if (latest < 3.0) {{
    html += `<div class="alert alert-warn">ROAS actual: ${{latest.toFixed(2)}}x — cerca del umbral mínimo</div>`;
  }} else {{
    html += `<div class="alert alert-ok">ROAS actual: ${{latest.toFixed(2)}}x — saludable</div>`;
  }}

  if (latestAcos > 35) {{
    html += `<div class="alert alert-danger">ACOS actual: ${{latestAcos.toFixed(1)}}% — supera el 35%</div>`;
  }}

  if (latestTacos > 20) {{
    html += `<div class="alert alert-warn">TACoS actual: ${{latestTacos.toFixed(1)}}% — inversión publicitaria alta vs ventas totales</div>`;
  }}

  let qw = D.qweeks[D.qweeks.length - 1];
  let month = MONTHS[MONTHS.length - 1];
  if (qw < 4) {{
    html += `<div class="alert alert-warn">${{month}} tiene solo ${{qw}} semana(s) de datos — los números parciales no son comparables</div>`;
  }}

  document.getElementById('alerts').innerHTML = html;
}}

// ── SUMMARY CARDS ───────────────────────────────────────────
function renderCards() {{
  const s = SUMMARY;
  const cards = [
    {{ label: 'Gross Sales', value: fmtFull(s.gsales_current), change: pct(s.gsales_current, s.gsales_prev), note: `${{s.qweeks_current}} semanas` }},
    {{ label: 'Net Profit', value: fmtFull(s.profit_current), change: pct(s.profit_current, s.profit_prev) }},
    {{ label: 'ROAS', value: s.roas_current.toFixed(2) + 'x', change: pct(s.roas_current, s.roas_prev) }},
    {{ label: 'Unidades', value: s.qty_current.toLocaleString(), change: pct(s.qty_current, s.qty_prev) }},
    {{ label: 'PPC Spend', value: fmtFull(s.ppc_spend_current), change: pct(s.ppc_spend_current, s.ppc_spend_prev) }},
    {{ label: 'TACoS', value: s.tacos_current.toFixed(1) + '%', change: pct(s.tacos_current, s.tacos_prev), invert: true }},
  ];

  let html = '';
  cards.forEach(c => {{
    let cls = c.change.cls;
    if (c.invert) cls = c.change.val > 0 ? 'down' : 'up';
    html += `<div class="card">
      <div class="card-label">${{c.label}}</div>
      <div class="card-value">${{c.value}}</div>
      <div class="card-change ${{cls}}">vs mes anterior: ${{c.change.txt}}</div>
      ${{c.note ? `<div class="card-note">${{c.note}}</div>` : ''}}
    </div>`;
  }});
  document.getElementById('summary-cards').innerHTML = html;
}}

// ── CHARTS ──────────────────────────────────────────────────
function chartGSales() {{
  Plotly.newPlot('chart-gsales', [
    {{ x: MONTHS, y: D.gsales, type: 'bar', marker: {{ color: '#3b82f6' }}, name: 'Gross Sales',
       text: D.gsales.map(v => fmt(v)), textposition: 'outside', textfont: {{ size: 11, color: '#9ca3af' }} }},
  ], {{...layout_base, title: {{ text: 'Gross Sales por Mes', font: {{ size: 15, color: '#e4e4e7' }} }} }}, config);
}}

function chartProfitBreakdown() {{
  Plotly.newPlot('chart-profit-breakdown', [
    {{ x: MONTHS, y: D.profit, type: 'bar', name: 'Profit (antes PPC)', marker: {{ color: '#22c55e' }} }},
    {{ x: MONTHS, y: D.ppc_spend, type: 'bar', name: 'PPC Spend', marker: {{ color: '#ef4444' }} }},
    {{ x: MONTHS, y: D.real_net_profit, type: 'scatter', mode: 'lines+markers', name: 'Net Profit', line: {{ color: '#eab308', width: 3 }}, marker: {{ size: 8 }} }},
  ], {{...layout_base, barmode: 'group', title: {{ text: 'Profit vs PPC Spend', font: {{ size: 15, color: '#e4e4e7' }} }} }}, config);
}}

function chartQty() {{
  Plotly.newPlot('chart-qty', [
    {{ x: MONTHS, y: D.qty, type: 'bar', marker: {{ color: '#a855f7' }}, name: 'Unidades',
       text: D.qty.map(v => v.toLocaleString()), textposition: 'outside', textfont: {{ size: 11, color: '#9ca3af' }} }},
  ], {{...layout_base, title: {{ text: 'Unidades Vendidas', font: {{ size: 15, color: '#e4e4e7' }} }} }}, config);
}}

function chartMargin() {{
  Plotly.newPlot('chart-margin', [
    {{ x: MONTHS, y: D.profit_margin, type: 'scatter', mode: 'lines+markers+text',
       line: {{ color: '#06b6d4', width: 3 }}, marker: {{ size: 8 }}, name: '% Margin',
       text: D.profit_margin.map(v => v.toFixed(1) + '%'), textposition: 'top center',
       textfont: {{ size: 11, color: '#9ca3af' }} }},
  ], {{...layout_base, title: {{ text: 'Profit Margin %', font: {{ size: 15, color: '#e4e4e7' }} }},
       yaxis: {{ ...layout_base.yaxis, ticksuffix: '%' }} }}, config);
}}

function chartPPC() {{
  Plotly.newPlot('chart-ppc', [
    {{ x: MONTHS, y: D.ppc_spend, type: 'bar', name: 'PPC Spend', marker: {{ color: '#f97316' }} }},
    {{ x: MONTHS, y: D.ppc_sales, type: 'bar', name: 'PPC Sales', marker: {{ color: '#3b82f6' }} }},
  ], {{...layout_base, barmode: 'group', title: {{ text: 'PPC Spend vs PPC Sales', font: {{ size: 15, color: '#e4e4e7' }} }} }}, config);
}}

function chartRoasAcos() {{
  Plotly.newPlot('chart-roas-acos', [
    {{ x: MONTHS, y: D.roas, type: 'scatter', mode: 'lines+markers+text', name: 'ROAS',
       line: {{ color: '#22c55e', width: 3 }}, marker: {{ size: 8 }},
       text: D.roas.map(v => v.toFixed(2) + 'x'), textposition: 'top center',
       textfont: {{ size: 11, color: '#9ca3af' }} }},
    {{ x: MONTHS, y: D.acos, type: 'scatter', mode: 'lines+markers', name: 'ACOS %',
       line: {{ color: '#ef4444', width: 2, dash: 'dash' }}, marker: {{ size: 6 }}, yaxis: 'y2' }},
  ], {{...layout_base,
       title: {{ text: 'ROAS & ACOS', font: {{ size: 15, color: '#e4e4e7' }} }},
       yaxis: {{ ...layout_base.yaxis, title: 'ROAS' }},
       yaxis2: {{ overlaying: 'y', side: 'right', gridcolor: gridColor, title: 'ACOS %', ticksuffix: '%' }},
       shapes: [{{ type: 'line', x0: MONTHS[0], x1: MONTHS[MONTHS.length-1], y0: 2.5, y1: 2.5,
                   line: {{ color: '#ef4444', width: 1, dash: 'dot' }} }}],
  }}, config);
}}

function chartTacos() {{
  Plotly.newPlot('chart-tacos', [
    {{ x: MONTHS, y: D.tacos, type: 'scatter', mode: 'lines+markers+text',
       line: {{ color: '#f97316', width: 3 }}, marker: {{ size: 8 }}, name: 'TACoS %',
       text: D.tacos.map(v => v.toFixed(1) + '%'), textposition: 'top center',
       textfont: {{ size: 11, color: '#9ca3af' }} }},
  ], {{...layout_base, title: {{ text: 'TACoS (Total Ad Cost of Sales)', font: {{ size: 15, color: '#e4e4e7' }} }},
       yaxis: {{ ...layout_base.yaxis, ticksuffix: '%' }} }}, config);
}}

function chartCpcVsTotal() {{
  Plotly.newPlot('chart-cpc-vs-total', [
    {{ x: MONTHS, y: D.cpc_vs_total, type: 'scatter', mode: 'lines+markers+text',
       line: {{ color: '#a855f7', width: 3 }}, marker: {{ size: 8 }}, name: 'PPC Sales / Total Sales %',
       text: D.cpc_vs_total.map(v => v.toFixed(1) + '%'), textposition: 'top center',
       textfont: {{ size: 11, color: '#9ca3af' }} }},
  ], {{...layout_base, title: {{ text: '% Ventas por PPC vs Total', font: {{ size: 15, color: '#e4e4e7' }} }},
       yaxis: {{ ...layout_base.yaxis, ticksuffix: '%' }} }}, config);
}}

function chartWeeklyProfit() {{
  Plotly.newPlot('chart-weekly-profit', [
    {{ x: MONTHS, y: D.profit_per_week, type: 'bar', marker: {{ color: '#22c55e' }}, name: 'Profit/Semana',
       text: D.profit_per_week.map(v => fmt(v)), textposition: 'outside',
       textfont: {{ size: 11, color: '#9ca3af' }} }},
  ], {{...layout_base, title: {{ text: 'Net Profit por Semana', font: {{ size: 15, color: '#e4e4e7' }} }} }}, config);
}}

function chartWeeklyQty() {{
  Plotly.newPlot('chart-weekly-qty', [
    {{ x: MONTHS, y: D.sales_per_week, type: 'bar', marker: {{ color: '#a855f7' }}, name: 'Unidades/Semana',
       text: D.sales_per_week.map(v => v.toFixed(0)), textposition: 'outside',
       textfont: {{ size: 11, color: '#9ca3af' }} }},
  ], {{...layout_base, title: {{ text: 'Unidades por Semana', font: {{ size: 15, color: '#e4e4e7' }} }} }}, config);
}}

function chartWeeklyPct() {{
  let dPct = D.dollar_weekly_pct.map(v => v === null ? 0 : v);
  let qPct = D.qty_weekly_pct.map(v => v === null ? 0 : v);
  let colors_d = dPct.map(v => v >= 0 ? '#22c55e' : '#ef4444');
  let colors_q = qPct.map(v => v >= 0 ? '#3b82f6' : '#f97316');

  Plotly.newPlot('chart-weekly-pct', [
    {{ x: MONTHS, y: dPct, type: 'bar', marker: {{ color: colors_d }}, name: '$ Profit/Sem %',
       text: dPct.map(v => (v >= 0 ? '+' : '') + v.toFixed(1) + '%'), textposition: 'outside',
       textfont: {{ size: 10, color: '#9ca3af' }} }},
    {{ x: MONTHS, y: qPct, type: 'scatter', mode: 'lines+markers', name: 'Qty/Sem %',
       line: {{ color: '#3b82f6', width: 2 }}, marker: {{ size: 6 }} }},
  ], {{...layout_base, title: {{ text: 'Cambio % Semanal (MoM)', font: {{ size: 15, color: '#e4e4e7' }} }},
       yaxis: {{ ...layout_base.yaxis, ticksuffix: '%', zeroline: true, zerolinecolor: '#4a4d5a' }} }}, config);
}}

function chartNetProfit() {{
  let cumulative = [];
  let sum = 0;
  D.real_net_profit.forEach(v => {{ sum += v; cumulative.push(Math.round(sum)); }});

  Plotly.newPlot('chart-net-profit', [
    {{ x: MONTHS, y: D.real_net_profit, type: 'bar', marker: {{ color: '#eab308' }}, name: 'Net Profit Mensual' }},
    {{ x: MONTHS, y: cumulative, type: 'scatter', mode: 'lines+markers', name: 'Acumulado',
       line: {{ color: '#06b6d4', width: 3 }}, marker: {{ size: 7 }}, yaxis: 'y2' }},
  ], {{...layout_base,
       title: {{ text: 'Net Profit & Acumulado', font: {{ size: 15, color: '#e4e4e7' }} }},
       yaxis2: {{ overlaying: 'y', side: 'right', gridcolor: 'transparent' }},
  }}, config);
}}


// ── PRODUCT SECTION ─────────────────────────────────────────
let currentMetric = 'sales';

function renderProductTabs() {{
  const metrics = [
    {{ key: 'sales', label: 'Ventas $' }},
    {{ key: 'profit', label: 'Profit $' }},
    {{ key: 'qty', label: 'Unidades' }},
  ];
  let html = '';
  metrics.forEach(m => {{
    html += `<button class="tab ${{m.key === currentMetric ? 'active' : ''}}" onclick="switchMetric('${{m.key}}')">${{m.label}}</button>`;
  }});
  document.getElementById('product-tabs').innerHTML = html;
}}

function switchMetric(metric) {{
  currentMetric = metric;
  renderProductTabs();
  chartProductTrend();
}}

function chartProductTrend() {{
  const top10 = PROD_TOTALS.slice(0, 10);
  const colors = ['#3b82f6','#22c55e','#a855f7','#f97316','#06b6d4','#eab308','#ef4444','#ec4899','#14b8a6','#8b5cf6'];
  let traces = [];
  top10.forEach((p, i) => {{
    let vals = PROD_DATA[p.name][currentMetric];
    traces.push({{
      x: MONTHS, y: vals, type: 'scatter', mode: 'lines+markers',
      name: p.name, line: {{ color: colors[i % colors.length], width: 2 }},
      marker: {{ size: 5 }},
    }});
  }});

  let titleMap = {{ sales: 'Ventas $ por Producto (Top 10)', profit: 'Profit $ por Producto (Top 10)', qty: 'Unidades por Producto (Top 10)' }};
  Plotly.newPlot('chart-product-trend', traces, {{
    ...layout_base,
    title: {{ text: titleMap[currentMetric], font: {{ size: 15, color: '#e4e4e7' }} }},
    legend: {{ ...layout_base.legend, y: -0.3 }},
  }}, config);
}}

function renderProductTable() {{
  let html = '<table><thead><tr><th>Producto</th><th>Ventas</th><th>Profit</th><th>Unidades</th><th>Margin %</th><th>Trend</th></tr></thead><tbody>';
  PROD_TOTALS.forEach(p => {{
    let pillCls = p.margin > 60 ? 'pill-green' : p.margin > 40 ? 'pill-yellow' : 'pill-red';
    let lastIdx = MONTHS.length - 1;
    let prevIdx = MONTHS.length - 2;
    let currSales = PROD_DATA[p.name].sales[lastIdx];
    let prevSales = PROD_DATA[p.name].sales[prevIdx];
    let trend = prevSales > 0 ? ((currSales - prevSales) / prevSales * 100) : 0;
    let trendCls = trend >= 0 ? 'pill-green' : 'pill-red';
    let trendTxt = (trend >= 0 ? '+' : '') + trend.toFixed(1) + '%';

    html += `<tr>
      <td>${{p.name}}</td>
      <td>${{fmtFull(p.sales)}}</td>
      <td>${{fmtFull(p.profit)}}</td>
      <td>${{p.qty.toLocaleString()}}</td>
      <td><span class="pill ${{pillCls}}">${{p.margin.toFixed(1)}}%</span></td>
      <td><span class="pill ${{trendCls}}">${{trendTxt}}</span></td>
    </tr>`;
  }});
  html += '</tbody></table>';
  document.getElementById('product-table-container').innerHTML = html;
}}


// ── SUMMARY TABLE ───────────────────────────────────────────
function renderSummaryTable() {{
  const rows = [
    {{ label: 'Semanas', key: 'qweeks', fmt: v => v }},
    {{ label: 'Unidades', key: 'qty', fmt: v => v.toLocaleString() }},
    {{ label: 'Gross Sales', key: 'gsales', fmt: v => fmtFull(v) }},
    {{ label: 'Profit (pre-PPC)', key: 'profit', fmt: v => fmtFull(v) }},
    {{ label: 'COGS', key: 'cogs', fmt: v => fmtFull(v) }},
    {{ label: 'Amazon Fees', key: 'amz_fee', fmt: v => fmtFull(v) }},
    {{ label: 'PPC Sales', key: 'ppc_sales', fmt: v => fmtFull(v) }},
    {{ label: 'PPC Spend', key: 'ppc_spend', fmt: v => fmtFull(v) }},
    {{ label: 'ROAS', key: 'roas', fmt: v => v.toFixed(2) + 'x' }},
    {{ label: 'ACOS', key: 'acos', fmt: v => v.toFixed(1) + '%' }},
    {{ label: 'TACoS', key: 'tacos', fmt: v => v.toFixed(1) + '%' }},
    {{ label: 'PPC/Total %', key: 'cpc_vs_total', fmt: v => v.toFixed(1) + '%' }},
    {{ label: 'Net Profit', key: 'real_net_profit', fmt: v => fmtFull(v) }},
    {{ label: 'Profit Margin %', key: 'profit_margin', fmt: v => v.toFixed(1) + '%' }},
    {{ label: 'Profit/Semana', key: 'profit_per_week', fmt: v => fmtFull(v) }},
    {{ label: 'Unid./Semana', key: 'sales_per_week', fmt: v => v.toFixed(1) }},
    {{ label: '$ Sem. %', key: 'dollar_weekly_pct', fmt: v => v === null ? '—' : (v >= 0 ? '+' : '') + v.toFixed(1) + '%' }},
    {{ label: 'Qty Sem. %', key: 'qty_weekly_pct', fmt: v => v === null ? '—' : (v >= 0 ? '+' : '') + v.toFixed(1) + '%' }},
  ];

  let html = '<table><thead><tr><th>Métrica</th>';
  MONTHS.slice().reverse().forEach(m => {{ html += `<th>${{m}}</th>`; }});
  html += '</tr></thead><tbody>';

  rows.forEach(r => {{
    html += `<tr><td>${{r.label}}</td>`;
    let vals = D[r.key].slice().reverse();
    vals.forEach(v => {{ html += `<td>${{r.fmt(v)}}</td>`; }});
    html += '</tr>';
  }});

  html += '</tbody></table>';
  document.getElementById('summary-table').innerHTML = html;
}}


// ── INIT ────────────────────────────────────────────────────
renderAlerts();
renderCards();
chartGSales();
chartProfitBreakdown();
chartQty();
chartMargin();
chartPPC();
chartRoasAcos();
chartTacos();
chartCpcVsTotal();
chartWeeklyProfit();
chartWeeklyQty();
chartWeeklyPct();
chartNetProfit();
renderProductTabs();
chartProductTrend();
renderProductTable();
renderSummaryTable();
</script>
</body>
</html>
"""

with open(OUTPUT, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"Dashboard saved to {OUTPUT}")
print(f"Months: {months_labels}")
print(f"Total GSales: ${sum(data['gsales']):,.2f}")
print(f"Total Net Profit: ${sum(data['real_net_profit']):,.2f}")
